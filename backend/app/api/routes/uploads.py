from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Form
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
import csv
import io

from app.db.session import get_db_session
from app.models.upload import Upload
from app.models.job import Job
from app.models.activity import Activity, ActivityType

router = APIRouter(prefix="/uploads", tags=["uploads"]) 


def _map_category_to_activity_type(category: str) -> ActivityType:
    mapping = {
        "VERKAUFSFOERDERUNG": ActivityType.sales,
        "IMAGE": ActivityType.branding,
        "EMPLOYER_BRANDING": ActivityType.employer_branding,
        "KUNDENPFLEGE": ActivityType.kundenpflege,
    }
    return mapping.get((category or "").upper(), ActivityType.sales)


def _parse_csv_to_activities(content: bytes) -> List[Dict[str, Any]]:
    text = content.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, Any]] = []
    for row in reader:
        rows.append(row)
    return rows


@router.get("")
def list_uploads(db: Session = Depends(get_db_session)):
    items = (
        db.query(Upload)
        .order_by(Upload.created_at.desc())
        .all()
    )
    return {
        "items": [
            {
                "id": str(u.id),
                "original_name": u.original_name,
                "file_type": u.file_type,
                "file_size": int(u.file_size or 0),
                "created_at": u.created_at,
            }
            for u in items
        ]
    }


@router.post("")
def upload_file(
    file: UploadFile = File(...),
    mapping: Optional[str] = Form(default=None),
    db: Session = Depends(get_db_session),
):
    """
    Accept CSV/Excel file and import rows as Activities.
    Columns supported (case-insensitive):
    title, category|type, status, weight, budget|budgetCHF, notes,
    start|start_date, end|end_date
    """
    # Save upload metadata
    upload = Upload(
        original_name=file.filename or "file",
        file_type=file.content_type or "",
        file_size=0,
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)

    # Read all content
    content = file.file.read()
    upload.file_size = len(content or b"")

    created_count = 0
    skipped_count = 0

    # Parse CSV or Excel
    rows: List[Dict[str, Any]] = []
    ctype = (file.content_type or "").lower()
    name_lower = (file.filename or "").lower()

    try:
        if ctype in {"text/csv", "application/csv", "text/plain"} or name_lower.endswith(".csv"):
            rows = _parse_csv_to_activities(content)
        elif name_lower.endswith(".xlsx") or ctype in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}:
            try:
                import openpyxl  # type: ignore
            except Exception:
                raise HTTPException(status_code=415, detail="Excel (.xlsx) not supported without openpyxl. Please upload CSV.")

            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(v or "").strip() for v in first]
            for r in ws.iter_rows(min_row=2, values_only=True):
                row = {headers[i].strip(): (r[i] if i is not None and i < len(r) else None) for i in range(len(headers))}
                rows.append(row)
        else:
            raise HTTPException(status_code=415, detail="Unsupported file type. Upload CSV or XLSX")

        # Normalize header names
        def g(row: Dict[str, Any], *keys: str):
            for k in keys:
                for variant in (k, k.lower(), k.upper(), k.replace(" ", "_")):
                    if variant in row and row[variant] not in (None, ""):
                        return row[variant]
            return None

        from datetime import datetime, date

        def parse_date(v):
            if v in (None, ""): return None
            if isinstance(v, (datetime, date)): return v if isinstance(v, date) else v.date()
            s = str(v).strip()
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
            try:
                # ISO
                return datetime.fromisoformat(s.replace('Z', '+00:00')).date()
            except Exception:
                return None

        # Optional mapping from form JSON { field -> headerName }
        mapping_dict: Dict[str, Any] = {}
        if mapping:
            try:
                import json as _json
                mapping_dict = _json.loads(mapping)
            except Exception:
                mapping_dict = {}

        def gm(row: Dict[str, Any], field: str, *fallbacks: str):
            header = (mapping_dict.get(field) or '').strip()
            if header:
                v = row.get(header)
                if v not in (None, ""):
                    return v
            return g(row, *fallbacks)

        for row in rows:
            try:
                title = gm(row, "title", "title", "name") or "Untitled"
                category = gm(row, "category", "category", "type") or "VERKAUFSFOERDERUNG"
                status = (gm(row, "status", "status") or "ACTIVE")
                budget = gm(row, "budget", "budget", "budgetCHF")
                weight = gm(row, "weight", "weight")
                notes = gm(row, "notes", "notes", "expected_output")
                start = parse_date(gm(row, "start", "start", "start_date"))
                end = parse_date(gm(row, "end", "end", "end_date"))

                activity = Activity(
                    title=str(title),
                    type=_map_category_to_activity_type(str(category)),
                    status=str(status).upper(),
                    budget=float(budget) if budget not in (None, "") else None,
                    weight=float(weight) if weight not in (None, "") else None,
                    expected_output=str(notes) if notes not in (None, "") else None,
                    start_date=start,
                    end_date=end,
                )
                db.add(activity)
                created_count += 1
            except Exception:
                skipped_count += 1

        db.commit()

        # Record a completed job
        rq_id = f"local-{upload.id}"
        job = Job(rq_id=rq_id, type="import_activities", status="finished", result=f"created={created_count};skipped={skipped_count}")
        db.add(job)
        db.commit()

        db.refresh(upload)
        return {
            "ok": True,
            "item": {
                "id": str(upload.id),
                "original_name": upload.original_name,
                "file_type": upload.file_type,
                "file_size": int(upload.file_size or 0),
                "created_at": upload.created_at,
            },
            "import": {"created": created_count, "skipped": skipped_count},
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/preview")
def preview_upload(file: UploadFile = File(...)):
    """Return headers, first rows and suggested mapping for CSV/XLSX."""
    content = file.file.read()
    ctype = (file.content_type or "").lower()
    name_lower = (file.filename or "").lower()

    rows: List[Dict[str, Any]] = []
    headers: List[str] = []
    if ctype in {"text/csv", "application/csv", "text/plain"} or name_lower.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= 4:
                break
    elif name_lower.endswith(".xlsx") or ctype in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}:
        try:
            import openpyxl  # type: ignore
        except Exception:
            raise HTTPException(status_code=415, detail="Excel (.xlsx) not supported without openpyxl. Please upload CSV.")
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(v or "").strip() for v in first]
        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            row = {headers[i2].strip(): (r[i2] if i2 is not None and i2 < len(r) else None) for i2 in range(len(headers))}
            rows.append(row)
            if i >= 4:
                break
    else:
        raise HTTPException(status_code=415, detail="Unsupported file type. Upload CSV or XLSX")

    # suggested mapping
    headers_l = [h.lower().replace(" ", "_") for h in headers]
    def suggest(*names: str) -> Optional[str]:
        for n in names:
            if n in headers_l:
                idx = headers_l.index(n)
                return headers[idx]
        return None

    suggested = {
        "title": suggest("title", "name"),
        "category": suggest("category", "type"),
        "status": suggest("status"),
        "budget": suggest("budget", "budgetchf"),
        "notes": suggest("notes", "expected_output"),
        "start": suggest("start", "start_date"),
        "end": suggest("end", "end_date"),
        "weight": suggest("weight"),
    }

    return {"headers": headers, "samples": rows, "suggested_mapping": suggested}


@router.get("/template.csv")
def template_csv() -> Response:
    headers = ["Title","Category","Status","BudgetCHF","Start","End","Notes","Weight"]
    example1 = ["Sommer Aktion","VERKAUFSFOERDERUNG","ACTIVE","1200","2025-06-01","2025-06-30","Promo Juni","1"]
    example2 = ["Brand Kampagne","IMAGE","PLANNED","5000","01.07.2025","31.07.2025","Awareness","2"]
    lines = [",".join(headers), ",".join(map(str, example1)), ",".join(map(str, example2))]
    content = "\n".join(lines)
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=activities-template.csv"},
    )


@router.get("/template.xlsx")
def template_xlsx() -> StreamingResponse:
    try:
        import openpyxl  # type: ignore
    except Exception:
        # Fallback: serve CSV if xlsx unsupported
        return StreamingResponse(
            iter([b"Title,Category,Status,BudgetCHF,Start,End,Notes,Weight\nSommer Aktion,VERKAUFSFOERDERUNG,ACTIVE,1200,2025-06-01,2025-06-30,Promo Juni,1\nBrand Kampagne,IMAGE,PLANNED,5000,01.07.2025,31.07.2025,Awareness,2\n"]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=activities-template.csv"},
        )

    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Title","Category","Status","BudgetCHF","Start","End","Notes","Weight"])
    ws.append(["Sommer Aktion","VERKAUFSFOERDERUNG","ACTIVE",1200,"2025-06-01","2025-06-30","Promo Juni",1])
    ws.append(["Brand Kampagne","IMAGE","PLANNED",5000,"01.07.2025","31.07.2025","Awareness",2])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=activities-template.xlsx"},
    )


